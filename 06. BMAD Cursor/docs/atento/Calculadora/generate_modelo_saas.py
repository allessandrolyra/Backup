import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
hdr_font = Font(bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
phase_font = Font(bold=True, color="FFFFFF", size=10)
phase_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
sub_font = Font(bold=True, size=10)
sub_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
tot_font_w = Font(bold=True, size=11, color="FFFFFF")
tot_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
title_font = Font(bold=True, size=14)
title2_font = Font(bold=True, size=11, color="2F5496")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
rec_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
rec_font = Font(bold=True, size=10, color="375623")
wrap = Alignment(vertical='center', wrap_text=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
right_a = Alignment(horizontal='right', vertical='center')

TAXA = 5.80
DURACAO_CHAMADA_S = 15
HORAS_DIA = 8
DIAS_MES = 22
CUSTO_SPEECH_HORA = 1.0  # USD per hour of audio
CUSTO_OPERACAO_MES = 800  # USD - custo analista parcial sustentacao
MARGEM_ALVO = 0.35  # 35% margem minima sobre preco de venda

def set_headers(ws, row, headers, widths=None):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = thin
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def write_row(ws, row, values, fmt=None, font=None, fill=None):
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.border = thin
        c.alignment = wrap
        if fmt and col in fmt:
            c.number_format = fmt[col]
            c.alignment = right_a
        if font:
            c.font = font
        if fill:
            c.fill = fill

# Cenarios de volume
cenarios = [
    {"nome": "MVP/PoC (10/min)", "calls_min": 10, "custo_fixo": 423.28, "custo_speech": 143.65, "custo_total": 566.93},
    {"nome": "Producao Inicial (50/min)", "calls_min": 50, "custo_fixo": 500, "custo_speech": 718.25, "custo_total": 1218.25},
    {"nome": "Producao (100/min)", "calls_min": 100, "custo_fixo": 978.54, "custo_speech": 4423.65, "custo_total": 5402.19},
    {"nome": "Escala (1.000/min)", "calls_min": 1000, "custo_fixo": 4259.84, "custo_speech": 44023.65, "custo_total": 48283.49},
]

for c in cenarios:
    c["calls_mes"] = c["calls_min"] * 60 * HORAS_DIA * DIAS_MES
    c["horas_audio"] = c["calls_mes"] * DURACAO_CHAMADA_S / 3600
    c["custo_por_chamada"] = c["custo_total"] / c["calls_mes"] if c["calls_mes"] > 0 else 0
    c["custo_fixo_por_chamada"] = c["custo_fixo"] / c["calls_mes"] if c["calls_mes"] > 0 else 0
    c["custo_var_por_chamada"] = CUSTO_SPEECH_HORA * DURACAO_CHAMADA_S / 3600


# ============================================================
# ABA 1: UNIT ECONOMICS
# ============================================================
ws1 = wb.active
ws1.title = "Unit Economics"

ws1.merge_cells('A1:H1')
ws1['A1'] = "Unit Economics — Custo por Chamada ASR"
ws1['A1'].font = title_font

ws1.merge_cells('A2:H2')
ws1['A2'] = "Decomposicao de custo fixo + variavel por cenario de volume | Chamada media: 15 segundos"
ws1['A2'].font = title2_font

set_headers(ws1, 4, [
    "Cenario", "Chamadas/min", "Chamadas/mes",
    "Custo Fixo (USD/mes)", "Custo Variavel Speech (USD/mes)", "Custo Total Azure (USD/mes)",
    "Custo por Chamada (USD)", "Custo por Chamada (BRL)"
], [28, 16, 18, 22, 26, 24, 22, 22])

for i, c in enumerate(cenarios, 5):
    write_row(ws1, i, [
        c["nome"], c["calls_min"], c["calls_mes"],
        c["custo_fixo"], c["custo_speech"], c["custo_total"],
        c["custo_por_chamada"], c["custo_por_chamada"] * TAXA
    ], fmt={3: '#,##0', 4: '$#,##0.00', 5: '$#,##0.00', 6: '$#,##0.00', 7: '$0.0000', 8: 'R$ 0.0000'})

r = 5 + len(cenarios) + 1
ws1.cell(row=r, column=1, value="Premissas:").font = Font(bold=True)
ws1.cell(row=r+1, column=1, value="- Chamada media: 15 segundos de audio")
ws1.cell(row=r+2, column=1, value="- Operacao: 8h/dia x 22 dias/mes")
ws1.cell(row=r+3, column=1, value="- Custo Speech STT: $1.00 por hora de audio processado")
ws1.cell(row=r+4, column=1, value="- Custo variavel por chamada (Speech): $0.00417 (fixo independente do volume)")
ws1.cell(row=r+5, column=1, value="- Custo fixo se dilui com volume (economia de escala)")
ws1.cell(row=r+6, column=1, value="- NAO inclui custo de operacao/sustentacao Foursys (~$800/mes)")


# ============================================================
# ABA 2: MODELO PAY-PER-CALL
# ============================================================
ws2 = wb.create_sheet("Modelo 1 - Pay-Per-Call")

ws2.merge_cells('A1:I1')
ws2['A1'] = "Modelo 1: Pay-Per-Call (Cobranca por Chamada Processada)"
ws2['A1'].font = title_font

ws2.merge_cells('A2:I2')
ws2['A2'] = "Cliente paga apenas pelo que usa. Sem mensalidade fixa."
ws2['A2'].font = title2_font

precos_chamada_brl = [0.08, 0.10, 0.15, 0.20, 0.25, 0.30]

set_headers(ws2, 4, [
    "Cenario", "Chamadas/mes", "Custo Azure (USD)", "Custo Operacao (USD)",
    "Custo Total Foursys (USD)", "Preco/chamada (BRL)", "Receita Mensal (BRL)",
    "Margem Bruta (BRL)", "Margem %"
], [28, 16, 18, 18, 22, 18, 20, 20, 12])

row = 5
for preco in precos_chamada_brl:
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    c = ws2.cell(row=row, column=1, value=f"Preco por chamada: R$ {preco:.2f}")
    c.font = phase_font
    c.fill = phase_fill
    c.border = thin
    for col in range(2, 10):
        ws2.cell(row=row, column=col).border = thin
    row += 1

    for cen in cenarios:
        custo_azure_brl = cen["custo_total"] * TAXA
        custo_op_brl = CUSTO_OPERACAO_MES * TAXA
        custo_total_brl = custo_azure_brl + custo_op_brl
        receita = cen["calls_mes"] * preco
        margem = receita - custo_total_brl
        margem_pct = margem / receita if receita > 0 else 0

        fill = green_fill if margem_pct > 0.3 else (yellow_fill if margem_pct > 0 else red_fill)

        values = [
            cen["nome"], cen["calls_mes"],
            cen["custo_total"], CUSTO_OPERACAO_MES,
            cen["custo_total"] + CUSTO_OPERACAO_MES,
            preco, receita, margem, margem_pct
        ]
        for col, v in enumerate(values, 1):
            c = ws2.cell(row=row, column=col, value=v)
            c.border = thin
            c.alignment = right_a if col > 1 else wrap
            if col == 2: c.number_format = '#,##0'
            elif col in (3, 4, 5): c.number_format = '$#,##0.00'
            elif col == 6: c.number_format = 'R$ 0.00'
            elif col in (7, 8): c.number_format = 'R$ #,##0.00'
            elif col == 9: c.number_format = '0.0%'
            if col == 9:
                c.fill = fill
        row += 1
    row += 1

r = row + 1
ws2.cell(row=r, column=1, value="Analise:").font = Font(bold=True)
ws2.cell(row=r+1, column=1, value="- Verde: margem > 30% (saudavel)")
ws2.cell(row=r+1, column=1).font = Font(color="375623")
ws2.cell(row=r+2, column=1, value="- Amarelo: margem 0-30% (baixa)")
ws2.cell(row=r+2, column=1).font = Font(color="9C6500")
ws2.cell(row=r+3, column=1, value="- Vermelho: prejuizo")
ws2.cell(row=r+3, column=1).font = Font(color="9C0006")
ws2.cell(row=r+4, column=1, value="- RISCO: Se volume for baixo, custos fixos nao se pagam")
ws2.cell(row=r+5, column=1, value="- Custo de operacao Foursys estimado: $800/mes (analista parcial + monitoramento)")


# ============================================================
# ABA 3: MODELO MENSALIDADE FIXA
# ============================================================
ws3 = wb.create_sheet("Modelo 2 - Mensalidade Fixa")

ws3.merge_cells('A1:H1')
ws3['A1'] = "Modelo 2: Mensalidade Fixa (Pacotes/Tiers)"
ws3['A1'].font = title_font

ws3.merge_cells('A2:H2')
ws3['A2'] = "Cliente escolhe um pacote mensal com chamadas incluidas."
ws3['A2'].font = title2_font

set_headers(ws3, 4, [
    "Tier / Pacote", "Chamadas Incluidas/mes", "Chamadas/min equiv.",
    "Custo Azure (USD/mes)", "Custo c/ Operacao (USD)",
    "Preco Sugerido (BRL/mes)", "Margem Bruta (BRL)", "Margem %"
], [25, 22, 20, 22, 22, 24, 20, 12])

tiers_fixo = []
for t in [
    {"nome": "Starter",      "calls": 50000,   "calls_min": 5,   "custo_fixo_infra": 423},
    {"nome": "Essential",    "calls": 160000,  "calls_min": 15,  "custo_fixo_infra": 500},
    {"nome": "Professional", "calls": 500000,  "calls_min": 50,  "custo_fixo_infra": 1000},
    {"nome": "Business",     "calls": 2000000, "calls_min": 150, "custo_fixo_infra": 2500},
    {"nome": "Enterprise",   "calls": 8000000, "calls_min": 500, "custo_fixo_infra": 8000},
]:
    speech = t["calls"] * DURACAO_CHAMADA_S / 3600 * CUSTO_SPEECH_HORA
    t["custo_azure"] = t["custo_fixo_infra"] + speech
    custo_total_usd = t["custo_azure"] + CUSTO_OPERACAO_MES
    t["preco_brl"] = round((custo_total_usd * TAXA) / (1 - MARGEM_ALVO) / 500) * 500
    tiers_fixo.append(t)
tiers = tiers_fixo

for i, t in enumerate(tiers, 5):
    custo_total = t["custo_azure"] + CUSTO_OPERACAO_MES
    custo_total_brl = custo_total * TAXA
    margem = t["preco_brl"] - custo_total_brl
    margem_pct = margem / t["preco_brl"] if t["preco_brl"] > 0 else 0

    fill = green_fill if margem_pct > 0.3 else (yellow_fill if margem_pct > 0 else red_fill)

    values = [
        t["nome"], t["calls"], t["calls_min"],
        t["custo_azure"], custo_total,
        t["preco_brl"], margem, margem_pct
    ]
    for col, v in enumerate(values, 1):
        c = ws3.cell(row=i, column=col, value=v)
        c.border = thin
        c.alignment = right_a if col > 1 else wrap
        if col == 2: c.number_format = '#,##0'
        elif col in (4, 5): c.number_format = '$#,##0.00'
        elif col in (6, 7): c.number_format = 'R$ #,##0.00'
        elif col == 8:
            c.number_format = '0.0%'
            c.fill = fill

r = 5 + len(tiers) + 2
ws3.cell(row=r, column=1, value="Vantagens:").font = Font(bold=True)
ws3.cell(row=r+1, column=1, value="- Receita previsivel para Foursys (MRR constante)")
ws3.cell(row=r+2, column=1, value="- Facil de orcar para o cliente (custo fixo mensal)")
ws3.cell(row=r+3, column=1, value="- Upgrade de tier conforme crescimento natural")
ws3.cell(row=r+5, column=1, value="Desvantagens:").font = Font(bold=True)
ws3.cell(row=r+6, column=1, value="- Cliente pode subutilizar o pacote")
ws3.cell(row=r+7, column=1, value="- Se exceder, precisa de regra de excedente ou upgrade")
ws3.cell(row=r+8, column=1, value="- Menos flexivel que pay-per-call")


# ============================================================
# ABA 4: MODELO HIBRIDO (RECOMENDADO)
# ============================================================
ws4 = wb.create_sheet("Modelo 3 - Hibrido (REC)")

ws4.merge_cells('A1:I1')
ws4['A1'] = "Modelo 3: Hibrido — Base Fixa + Excedente por Chamada (RECOMENDADO)"
ws4['A1'].font = title_font

ws4.merge_cells('A2:I2')
ws4['A2'] = "Mensalidade base cobre custos fixos + chamadas incluidas. Excedente cobrado por chamada."
ws4['A2'].font = title2_font

# Estrutura do modelo
ws4.cell(row=4, column=1, value="ESTRUTURA DO MODELO").font = Font(bold=True, size=11)

struct_headers = [
    "Tier", "Mensalidade Base (BRL)", "Chamadas Incluidas",
    "Preco Excedente (BRL/chamada)", "Chamadas/min equiv.",
    "Custo Azure Total (USD)", "Margem Base"
]
set_headers(ws4, 5, struct_headers, [22, 24, 22, 26, 20, 22, 22])

hibrido_tiers = [
    {"nome": "PoC",          "calls_incl":    50000, "excedente_brl": 0.12, "calls_min":   5, "custo_fixo_infra": 423},
    {"nome": "Starter",      "calls_incl":   160000, "excedente_brl": 0.10, "calls_min":  15, "custo_fixo_infra": 500},
    {"nome": "Professional", "calls_incl":   500000, "excedente_brl": 0.08, "calls_min":  50, "custo_fixo_infra": 1000},
    {"nome": "Business",     "calls_incl":  2000000, "excedente_brl": 0.06, "calls_min": 150, "custo_fixo_infra": 2500},
    {"nome": "Enterprise",   "calls_incl":  8000000, "excedente_brl": 0.04, "calls_min": 500, "custo_fixo_infra": 8000},
]

for t in hibrido_tiers:
    speech_incl = t["calls_incl"] * DURACAO_CHAMADA_S / 3600 * CUSTO_SPEECH_HORA
    t["custo_azure_total"] = t["custo_fixo_infra"] + speech_incl
    custo_total_usd = t["custo_azure_total"] + CUSTO_OPERACAO_MES
    custo_total_brl = custo_total_usd * TAXA
    t["custo_total_brl"] = custo_total_brl
    t["base_brl"] = round(custo_total_brl / (1 - MARGEM_ALVO) / 500) * 500

for i, t in enumerate(hibrido_tiers, 6):
    margem_base = t["base_brl"] - t["custo_total_brl"]
    margem_pct_base = margem_base / t["base_brl"] if t["base_brl"] > 0 else 0

    values = [
        t["nome"], t["base_brl"], t["calls_incl"],
        t["excedente_brl"], t["calls_min"],
        t["custo_azure_total"],
        f"{margem_pct_base:.0%} margem"
    ]
    for col, v in enumerate(values, 1):
        c = ws4.cell(row=i, column=col, value=v)
        c.border = thin
        c.alignment = right_a if col > 1 else wrap
        if col == 2: c.number_format = 'R$ #,##0.00'
        elif col == 3: c.number_format = '#,##0'
        elif col == 4: c.number_format = 'R$ 0.00'
        elif col == 6: c.number_format = '$#,##0.00'

# Simulacao de receita
sim_row = 6 + len(hibrido_tiers) + 2
ws4.cell(row=sim_row, column=1, value="SIMULACAO DE RECEITA POR CENARIO").font = Font(bold=True, size=11)

sim_headers = [
    "Tier", "Chamadas Reais/mes", "Chamadas Incluidas",
    "Excedente (chamadas)", "Receita Base (BRL)", "Receita Excedente (BRL)",
    "Receita Total (BRL)", "Custo Total Foursys (BRL)", "Margem (BRL)", "Margem %"
]
set_headers(ws4, sim_row + 1, sim_headers, [22, 20, 20, 18, 18, 20, 20, 22, 18, 10])

simulacoes = [
    {"tier": "PoC", "idx": 0, "calls_real": 30000},
    {"tier": "PoC", "idx": 0, "calls_real": 50000},
    {"tier": "PoC", "idx": 0, "calls_real": 80000},
    {"tier": "Starter", "idx": 1, "calls_real": 100000},
    {"tier": "Starter", "idx": 1, "calls_real": 160000},
    {"tier": "Starter", "idx": 1, "calls_real": 250000},
    {"tier": "Professional", "idx": 2, "calls_real": 300000},
    {"tier": "Professional", "idx": 2, "calls_real": 500000},
    {"tier": "Professional", "idx": 2, "calls_real": 800000},
    {"tier": "Business", "idx": 3, "calls_real": 1500000},
    {"tier": "Business", "idx": 3, "calls_real": 2000000},
    {"tier": "Business", "idx": 3, "calls_real": 3000000},
]

sr = sim_row + 2
for s in simulacoes:
    t = hibrido_tiers[s["idx"]]
    calls_real = s["calls_real"]
    calls_incl = t["calls_incl"]
    excedente_calls = max(0, calls_real - calls_incl)
    receita_base = t["base_brl"]
    receita_exc = excedente_calls * t["excedente_brl"]
    receita_total = receita_base + receita_exc

    horas_audio_real = calls_real * DURACAO_CHAMADA_S / 3600
    custo_speech_real = horas_audio_real * CUSTO_SPEECH_HORA
    custo_azure_real = t["custo_fixo_infra"] + custo_speech_real
    custo_total_brl = (custo_azure_real + CUSTO_OPERACAO_MES) * TAXA

    margem = receita_total - custo_total_brl
    margem_pct = margem / receita_total if receita_total > 0 else 0
    fill = green_fill if margem_pct > 0.3 else (yellow_fill if margem_pct > 0 else red_fill)

    values = [
        t["nome"], calls_real, calls_incl,
        excedente_calls, receita_base, receita_exc,
        receita_total, custo_total_brl, margem, margem_pct
    ]
    for col, v in enumerate(values, 1):
        c = ws4.cell(row=sr, column=col, value=v)
        c.border = thin
        c.alignment = right_a if col > 1 else wrap
        if col in (2, 3, 4): c.number_format = '#,##0'
        elif col in (5, 6, 7, 8, 9): c.number_format = 'R$ #,##0.00'
        elif col == 10:
            c.number_format = '0.0%'
            c.fill = fill
    sr += 1

# Justificativa
jr = sr + 2
ws4.cell(row=jr, column=1, value="POR QUE HIBRIDO E O MELHOR MODELO:").font = Font(bold=True, size=11)
justificativas = [
    "1. Protege Foursys: mensalidade base cobre custos fixos mesmo com volume baixo",
    "2. Justo para o cliente: paga menos se usar menos, excedente proporcional",
    "3. Incentiva crescimento: preco de excedente diminui nos tiers maiores",
    "4. Previsibilidade: ambos os lados sabem o custo minimo mensal",
    "5. Escalavel: upgrade de tier natural conforme volume cresce",
    "6. Competitivo: excedente de R$ 0.04-0.12/chamada e compativel com mercado STT",
]
for i, j in enumerate(justificativas):
    ws4.cell(row=jr + 1 + i, column=1, value=j)


# ============================================================
# ABA 5: MODELO POR MINUTO DE AUDIO
# ============================================================
ws5 = wb.create_sheet("Modelo 4 - Por Minuto Audio")

ws5.merge_cells('A1:G1')
ws5['A1'] = "Modelo 4: Cobranca por Minuto de Audio Processado"
ws5['A1'].font = title_font

ws5.merge_cells('A2:G2')
ws5['A2'] = "Modelo transparente baseado no tempo de audio. Similar a como Azure cobra da Foursys."
ws5['A2'].font = title2_font

set_headers(ws5, 4, [
    "Cenario", "Horas Audio/mes", "Custo Azure (USD/mes)",
    "Custo c/ Operacao (USD)", "Preco Sugerido (BRL/hora)",
    "Receita Mensal (BRL)", "Margem %"
], [28, 18, 20, 22, 22, 20, 12])

precos_hora_brl = [10, 12, 15]

row5 = 5
for preco_h in precos_hora_brl:
    ws5.merge_cells(start_row=row5, start_column=1, end_row=row5, end_column=7)
    c = ws5.cell(row=row5, column=1, value=f"Preco: R$ {preco_h:.0f} por hora de audio")
    c.font = phase_font
    c.fill = phase_fill
    c.border = thin
    for col in range(2, 8):
        ws5.cell(row=row5, column=col).border = thin
    row5 += 1

    for cen in cenarios:
        custo_total = (cen["custo_total"] + CUSTO_OPERACAO_MES) * TAXA
        receita = cen["horas_audio"] * preco_h
        margem_pct = (receita - custo_total) / receita if receita > 0 else 0
        fill = green_fill if margem_pct > 0.3 else (yellow_fill if margem_pct > 0 else red_fill)

        values = [
            cen["nome"], round(cen["horas_audio"], 1),
            cen["custo_total"], cen["custo_total"] + CUSTO_OPERACAO_MES,
            preco_h, receita, margem_pct
        ]
        for col, v in enumerate(values, 1):
            c = ws5.cell(row=row5, column=col, value=v)
            c.border = thin
            c.alignment = right_a if col > 1 else wrap
            if col in (3, 4): c.number_format = '$#,##0.00'
            elif col == 5: c.number_format = 'R$ #,##0.00'
            elif col == 6: c.number_format = 'R$ #,##0.00'
            elif col == 7:
                c.number_format = '0.0%'
                c.fill = fill
        row5 += 1
    row5 += 1

r5 = row5 + 1
ws5.cell(row=r5, column=1, value="Referencia de mercado:").font = Font(bold=True)
ws5.cell(row=r5+1, column=1, value="- Azure Speech STT: ~R$ 5.80/hora (custo Foursys)")
ws5.cell(row=r5+2, column=1, value="- Google Cloud STT: ~R$ 7.00/hora (preco publico)")
ws5.cell(row=r5+3, column=1, value="- AWS Transcribe: ~R$ 8.70/hora (preco publico)")
ws5.cell(row=r5+4, column=1, value="- Modelo NAO recomendado: cliente nao pensa em 'horas de audio'")


# ============================================================
# ABA 6: COMPARATIVO
# ============================================================
ws6 = wb.create_sheet("Comparativo e Recomendacao")

ws6.merge_cells('A1:G1')
ws6['A1'] = "Comparativo dos Modelos de Precificacao SaaS"
ws6['A1'].font = title_font

set_headers(ws6, 3, [
    "Criterio", "Pay-Per-Call", "Mensalidade Fixa", "Hibrido (RECOMENDADO)",
    "Por Minuto Audio"
], [32, 28, 28, 30, 28])

criterios = [
    ["Previsibilidade de receita (Foursys)", "Baixa", "Alta", "Media-Alta", "Baixa"],
    ["Previsibilidade de custo (Cliente)", "Alta (paga o que usa)", "Alta (valor fixo)", "Alta (base conhecida)", "Media"],
    ["Protecao contra volume baixo", "Nenhuma (risco alto)", "Total", "Mensalidade base cobre fixos", "Nenhuma"],
    ["Justica no pricing", "Alta", "Media (pode subutilizar)", "Alta", "Alta"],
    ["Facilidade de entendimento", "Muito facil", "Facil", "Facil", "Dificil (horas de audio?)"],
    ["Escalabilidade do modelo", "Linear", "Upgrades de tier", "Organica (base + excedente)", "Linear"],
    ["Incentivo ao crescimento", "Nenhum", "Desconto por volume", "Excedente decrescente", "Nenhum"],
    ["Competitividade no mercado", "Alta", "Media", "Alta", "Baixa"],
    ["Complexidade de billing", "Baixa", "Nenhuma", "Media", "Media"],
    ["Risco para Foursys", "ALTO (volume incerto)", "BAIXO", "BAIXO", "ALTO"],
    ["Risco para Cliente", "BAIXO", "MEDIO (desperdicar)", "BAIXO", "MEDIO"],
    ["Alinhamento com valor entregue", "Alto", "Medio", "Alto", "Baixo"],
]

for i, row_data in enumerate(criterios, 4):
    for col, v in enumerate(row_data, 1):
        c = ws6.cell(row=i, column=col, value=v)
        c.border = thin
        c.alignment = wrap
        if col == 4:
            c.fill = rec_fill
            c.font = rec_font

# Nota final
nr = 4 + len(criterios) + 2
ws6.merge_cells(start_row=nr, start_column=1, end_row=nr, end_column=5)
c = ws6.cell(row=nr, column=1, value="RECOMENDACAO: MODELO HIBRIDO (Base Fixa + Excedente)")
c.font = tot_font_w
c.fill = tot_fill
for col in range(2, 6):
    ws6.cell(row=nr, column=col).fill = tot_fill
    ws6.cell(row=nr, column=col).border = thin
c.border = thin

rec_texts = [
    "",
    "O modelo hibrido equilibra os interesses de ambas as partes:",
    "",
    "PARA FOURSYS:",
    "- Mensalidade base garante cobertura dos custos fixos (VPN, VM, APIM, Redis, Monitor)",
    "- Receita recorrente previsivel (MRR) facilita planejamento financeiro",
    "- Excedente gera upside quando cliente cresce",
    "- Risco de prejuizo praticamente eliminado",
    "",
    "PARA ATENTO:",
    "- Custo mensal previsivel e orcavel",
    "- Paga proporcionalmente ao uso real (excedente justo)",
    "- Incentivo claro de escala: tiers maiores tem excedente mais barato",
    "- Sem surpresas: sabe exatamente o custo base + custo marginal",
    "",
    "SUGESTAO DE CONTRATO:",
    "- Periodo minimo: 12 meses (para justificar setup)",
    "- Tier inicial: Starter (R$ 8.000/mes com 160.000 chamadas)",
    "- Revisao trimestral de tier conforme volume real",
    "- SLA: 99.5% uptime com creditos se descumprir",
    "- Excedente faturado mensalmente com base em relatorio de uso",
]

for i, t in enumerate(rec_texts):
    ws6.cell(row=nr + 1 + i, column=1, value=t)
    if t.startswith("PARA") or t.startswith("SUGESTAO"):
        ws6.cell(row=nr + 1 + i, column=1).font = Font(bold=True)

# Salvar
output = os.path.join(
    r'c:\01. Foursys\06. BMAD Cursor\docs\atento\Calculadora',
    '[Atento] Modelo Precificacao SaaS - ASR Cloud v2.xlsx'
)
wb.save(output)

print(f"Excel gerado: {output}")
print(f"\nAbas:")
print(f"  1. Unit Economics - custo por chamada por cenario")
print(f"  2. Modelo 1 - Pay-Per-Call (6 faixas de preco x 4 cenarios)")
print(f"  3. Modelo 2 - Mensalidade Fixa (5 tiers)")
print(f"  4. Modelo 3 - Hibrido RECOMENDADO (5 tiers + 12 simulacoes)")
print(f"  5. Modelo 4 - Por Minuto de Audio (3 precos x 4 cenarios)")
print(f"  6. Comparativo e Recomendacao Final")
