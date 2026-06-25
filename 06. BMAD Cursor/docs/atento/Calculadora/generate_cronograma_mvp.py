import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = Workbook()

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_font = Font(bold=True, color="FFFFFF", size=10)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
phase_font = Font(bold=True, size=10, color="FFFFFF")
phase_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
subtotal_font = Font(bold=True, size=10)
subtotal_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
total_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
total_font_white = Font(bold=True, size=11, color="FFFFFF")
title_font = Font(bold=True, size=14)
subtitle_font = Font(bold=True, size=11, color="2F5496")
wrap_align = Alignment(vertical='center', wrap_text=True)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ============================================================
# ABA 1: CRONOGRAMA POR RECURSO / ATIVIDADE
# ============================================================
ws = wb.active
ws.title = "Cronograma por Atividade"

ws.merge_cells('A1:H1')
ws['A1'] = "Atento ASR Cloud - Cronograma de Configuracao MVP PoC"
ws['A1'].font = title_font

ws.merge_cells('A2:H2')
ws['A2'] = "Horas por atividade, recurso Azure e perfil de analista | Maio 2026"
ws['A2'].font = subtitle_font

headers = [
    "Fase",
    "Recurso / Servico Azure",
    "Atividade",
    "Perfil Analista",
    "Horas Estimadas",
    "Dependencia",
    "Semana",
    "Observacoes"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

data = [
    # FASE 1: PLANEJAMENTO E DESIGN
    ["FASE 1: PLANEJAMENTO E DESIGN", None, None, None, None, None, None, None],
    [None, "Geral", "Kickoff tecnico e alinhamento de requisitos", "Arquiteto Cloud", 4, "-", "S1", "Reuniao com stakeholders Foursys + Atento"],
    [None, "Geral", "Revisao da arquitetura C4 e ajustes para PoC", "Arquiteto Cloud", 4, "-", "S1", "Simplificar para cenario MVP"],
    [None, "Geral", "Definicao do plano de enderecamento IP (VNet/subnets)", "Arquiteto Cloud", 2, "-", "S1", "Alinhar ranges com Atento (evitar conflito)"],
    [None, "Geral", "Levantamento de pre-requisitos com Atento (Avaya, IPs, firewall)", "Arquiteto Cloud", 4, "-", "S1", "Dependencia critica do cliente"],
    [None, "Geral", "Documentacao do plano de implementacao", "Arquiteto Cloud", 4, "-", "S1", "Checklist de recursos e ordem de deploy"],

    # FASE 2: PROVISIONAMENTO DE INFRA BASE
    ["FASE 2: PROVISIONAMENTO DE INFRAESTRUTURA BASE", None, None, None, None, None, None, None],
    [None, "Azure Subscription", "Criacao/config da subscription e Resource Groups", "Cloud Engineer", 2, "-", "S2", "Naming convention, tags, policies"],
    [None, "VNet + Subnets", "Criar VNet 10.100.0.0/16 com subnets segmentadas", "Cloud Engineer", 3, "Subscription", "S2", "GatewaySubnet, Subnet-APIM, Subnet-App, Subnet-PE"],
    [None, "NSGs", "Configurar Network Security Groups por subnet", "Cloud Engineer", 3, "VNet", "S2", "Rules: Allow Atento IPs, Deny all inbound"],
    [None, "Route Tables", "Configurar tabelas de roteamento (UDR)", "Cloud Engineer", 2, "VNet", "S2", "Rotas para trafego VPN"],
    [None, "Key Vault", "Provisionar Key Vault + Managed Identity", "Cloud Engineer", 2, "Subscription", "S2", "Soft-delete habilitado, access policies"],
    [None, "Key Vault", "Armazenar certificados VPN e API keys", "Cloud Engineer", 1, "Key Vault", "S2", "Gerar e armazenar secrets"],

    # FASE 3: CONECTIVIDADE VPN
    ["FASE 3: CONECTIVIDADE VPN", None, None, None, None, None, None, None],
    [None, "VPN Gateway", "Provisionar VPN Gateway VpnGw1AZ", "Cloud Engineer", 2, "VNet", "S2-S3", "Deploy demora ~30-45min; config apos"],
    [None, "VPN Gateway", "Configurar Local Network Gateway (IPs Atento)", "Cloud Engineer", 2, "VPN Gateway", "S3", "Depende dos IPs fornecidos pela Atento"],
    [None, "VPN Gateway", "Criar Connection S2S (IPSec/IKEv2)", "Cloud Engineer", 2, "Local GW", "S3", "Shared key, DH Group, cipher suite"],
    [None, "VPN Gateway", "Teste de conectividade tunel VPN com Atento", "Cloud Engineer", 4, "Connection", "S3", "Validar ping, traceroute, latencia"],
    [None, "VPN Gateway", "Troubleshooting e ajuste de parametros VPN", "Cloud Engineer", 4, "Teste VPN", "S3", "Buffer para problemas de firewall/NAT"],
    [None, "VPN Gateway", "Configurar BGP (se aplicavel)", "Arquiteto Cloud", 2, "Connection", "S3", "ASN, peering addresses"],

    # FASE 4: SERVIÇOS DE DADOS E SEGURANÇA
    ["FASE 4: SERVICOS DE DADOS E SEGURANCA", None, None, None, None, None, None, None],
    [None, "Redis Cache", "Provisionar Azure Cache for Redis Standard C0", "Cloud Engineer", 1, "VNet", "S3", "Standard para suporte a PE"],
    [None, "Redis Cache", "Configurar Private Endpoint do Redis", "Cloud Engineer", 1, "Redis", "S3", "DNS privado + NSG"],
    [None, "Redis Cache", "Configurar firewall rules e acesso VNet", "Cloud Engineer", 1, "Redis PE", "S3", "Bloquear acesso publico"],
    [None, "Private Endpoints", "Criar PE para Speech Services", "Cloud Engineer", 1, "Speech Svc", "S3", "Private DNS Zone + link"],
    [None, "Private Endpoints", "Validar resolucao DNS privada (todos PEs)", "Cloud Engineer", 2, "Todos PEs", "S3", "nslookup, teste de conectividade interna"],

    # FASE 5: API MANAGEMENT
    ["FASE 5: API MANAGEMENT", None, None, None, None, None, None, None],
    [None, "API Management", "Provisionar APIM Developer tier", "Cloud Engineer", 2, "VNet", "S3", "Deploy demora ~30-40min"],
    [None, "API Management", "Configurar VNet injection (internal mode)", "Cloud Engineer", 3, "APIM", "S3", "Subnet dedicada, NSG especifico"],
    [None, "API Management", "Criar API definitions e operacoes (WebSocket)", "Desenvolvedor Backend", 4, "APIM VNet", "S4", "Endpoint /asr/transcribe"],
    [None, "API Management", "Configurar policies (rate-limit, auth, CORS, logging)", "Desenvolvedor Backend", 3, "API defs", "S4", "API Key validation, IP filter"],
    [None, "API Management", "Testar fluxo APIM -> backend (VM)", "Desenvolvedor Backend", 2, "Policies", "S4", "Validar roteamento e auth"],

    # FASE 6: COMPUTE (VM)
    ["FASE 6: COMPUTE (VM)", None, None, None, None, None, None, None],
    [None, "Virtual Machine", "Provisionar VM D2s v3 (Linux) na subnet", "Cloud Engineer", 2, "VNet", "S3", "Ubuntu 22.04 LTS, SSH key"],
    [None, "Virtual Machine", "Hardening do OS (updates, firewall, SSH)", "Cloud Engineer", 2, "VM", "S3", "Desabilitar senha, apenas SSH key"],
    [None, "Virtual Machine", "Instalar runtime (.NET / Node.js) e dependencias", "Cloud Engineer", 2, "VM hardened", "S3", "SDK do Speech, runtime da app"],
    [None, "Virtual Machine", "Configurar Managed Identity e acesso ao Key Vault", "Cloud Engineer", 1, "VM + KV", "S3", "System-assigned identity"],
    [None, "Virtual Machine", "Configurar scripts de deploy e restart", "Cloud Engineer", 2, "Runtime", "S4", "Systemd service, auto-start"],

    # FASE 7: SPEECH SERVICES E CUSTOM MODEL
    ["FASE 7: SPEECH SERVICES E CUSTOM MODEL", None, None, None, None, None, None, None],
    [None, "Speech Services", "Provisionar Azure Speech Services (S0) em Brazil South", "Cloud Engineer", 1, "Subscription", "S2", "Verificar disponibilidade na regiao"],
    [None, "Speech Services", "Configurar Private Endpoint do Speech", "Cloud Engineer", 1, "Speech Svc", "S3", "Private DNS Zone"],
    [None, "Custom Speech Model", "Preparar dataset de treinamento (CNPJs alfanumericos)", "Desenvolvedor Backend", 8, "-", "S2-S3", "Gravar/sintetizar audios com variacoes foneticas"],
    [None, "Custom Speech Model", "Treinar Custom Speech Model pt-BR", "Desenvolvedor Backend", 4, "Dataset", "S3-S4", "Upload dataset, configurar treinamento"],
    [None, "Custom Speech Model", "Avaliar acuracia do modelo (WER/CER)", "Desenvolvedor Backend", 4, "Modelo treinado", "S4", "Testar com CNPJs reais e variantes"],
    [None, "Custom Speech Model", "Deploy do Custom endpoint", "Desenvolvedor Backend", 2, "Modelo aprovado", "S4", "Endpoint dedicado para inferencia"],
    [None, "Custom Speech Model", "Re-treino e ajustes (iteracao)", "Desenvolvedor Backend", 8, "Avaliacao", "S4-S5", "Buffer para melhorar acuracia"],

    # FASE 8: DESENVOLVIMENTO DA APLICAÇÃO
    ["FASE 8: DESENVOLVIMENTO DA APLICACAO ASR", None, None, None, None, None, None, None],
    [None, "ASR Orchestrator", "Desenvolver WebSocket Handler (receber audio stream)", "Desenvolvedor Backend", 12, "VM + APIM", "S3-S4", "ASP.NET Core ou Node.js"],
    [None, "ASR Orchestrator", "Desenvolver Speech Client (integracao Speech SDK)", "Desenvolvedor Backend", 8, "Speech Svc", "S4", "Real-time STT, resultados parciais/finais"],
    [None, "ASR Orchestrator", "Desenvolver Audio Preprocessor (G.711 -> PCM 16kHz)", "Desenvolvedor Backend", 6, "WS Handler", "S4", "Conversao de codec, VAD"],
    [None, "ASR Orchestrator", "Desenvolver CNPJ Parser (texto -> CNPJ alfanumerico)", "Desenvolvedor Backend", 8, "Speech Client", "S4-S5", "Heuristicas foneticas, normalizacao"],
    [None, "ASR Orchestrator", "Desenvolver CNPJ Validator (digitos verificadores)", "Desenvolvedor Backend", 4, "Parser", "S5", "Algoritmo novo CNPJ alfanumerico"],
    [None, "ASR Orchestrator", "Desenvolver Retry & Fallback Handler", "Desenvolvedor Backend", 4, "Validator", "S5", "Max 3 tentativas, fallback para operador"],
    [None, "ASR Orchestrator", "Desenvolver Response Builder (JSON estruturado)", "Desenvolvedor Backend", 2, "Retry", "S5", "CNPJ, confianca, status"],
    [None, "ASR Orchestrator", "Integrar App Insights SDK (telemetria)", "Desenvolvedor Backend", 3, "App base", "S5", "Metricas de latencia, erros, throughput"],
    [None, "ASR Orchestrator", "Integrar Redis para session cache", "Desenvolvedor Backend", 3, "Redis PE", "S5", "Sessao por call-id"],

    # FASE 9: OBSERVABILIDADE
    ["FASE 9: OBSERVABILIDADE", None, None, None, None, None, None, None],
    [None, "Azure Monitor", "Criar Log Analytics Workspace", "Cloud Engineer", 1, "Subscription", "S4", "Retencao 30 dias"],
    [None, "App Insights", "Configurar Application Insights + connection string", "Cloud Engineer", 1, "Log Analytics", "S4", "Linked ao workspace"],
    [None, "Azure Monitor", "Criar alertas criticos (VPN down, latencia, error rate)", "Cloud Engineer", 2, "App Insights", "S5", "Action group: email/SMS"],
    [None, "Azure Monitor", "Criar dashboard operacional basico", "Cloud Engineer", 2, "Metricas", "S5", "KPIs: latencia, taxa ASR, erros"],

    # FASE 10: INTEGRAÇÃO E TESTES
    ["FASE 10: INTEGRACAO E TESTES END-TO-END", None, None, None, None, None, None, None],
    [None, "Integracao Avaya", "Configurar SBC para enviar audio via VPN/WebSocket", "Arquiteto Cloud", 8, "VPN + App", "S5-S6", "Depende da equipe Atento"],
    [None, "Integracao Avaya", "Testar fluxo completo: URA -> VPN -> APIM -> ASR -> resposta", "Arquiteto Cloud", 6, "Integracao", "S6", "Ligacao real com CNPJ de teste"],
    [None, "Teste E2E", "Testes funcionais com diferentes CNPJs alfanumericos", "QA / Tester", 8, "Integracao OK", "S6", "Cenarios: sucesso, retry, fallback"],
    [None, "Teste E2E", "Testes de latencia (medir E2E < 1s)", "QA / Tester", 4, "Funcional OK", "S6", "Medir cada etapa do pipeline"],
    [None, "Teste E2E", "Testes com sotaques regionais e ruido de fundo", "QA / Tester", 4, "Funcional OK", "S6", "SP, RJ, NE, Sul"],
    [None, "Teste E2E", "Testes de resiliencia (derrubar componentes)", "QA / Tester", 4, "Funcional OK", "S6-S7", "VM restart, Redis flush, VPN drop"],
    [None, "Geral", "Documentacao tecnica e handover", "Arquiteto Cloud", 4, "Testes OK", "S7", "Runbook operacional, diagrama atualizado"],
    [None, "Geral", "Apresentacao de resultados e proximos passos", "Arquiteto Cloud", 2, "Doc", "S7", "Reuniao com Atento"],
]

row_idx = 5
total_hours = 0

for item in data:
    if item[0] and item[1] is None:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=8)
        cell = ws.cell(row=row_idx, column=1, value=item[0])
        cell.font = phase_font
        cell.fill = phase_fill
        cell.alignment = Alignment(vertical='center')
        cell.border = thin_border
        for c in range(2, 9):
            ws.cell(row=row_idx, column=c).border = thin_border
        row_idx += 1
    else:
        for col_idx, value in enumerate(item, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value else "")
            cell.border = thin_border
            cell.alignment = wrap_align
            if col_idx == 5 and value:
                cell.alignment = center_align
                total_hours += value
        row_idx += 1

# TOTAL GERAL
ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
cell = ws.cell(row=row_idx, column=1, value="TOTAL GERAL DE HORAS")
cell.font = total_font_white
cell.fill = total_fill
cell.border = thin_border
for c in range(2, 5):
    ws.cell(row=row_idx, column=c).fill = total_fill
    ws.cell(row=row_idx, column=c).border = thin_border

cell = ws.cell(row=row_idx, column=5, value=total_hours)
cell.font = total_font_white
cell.fill = total_fill
cell.alignment = center_align
cell.border = thin_border

for c in range(6, 9):
    ws.cell(row=row_idx, column=c).fill = total_fill
    ws.cell(row=row_idx, column=c).border = thin_border

col_widths_1 = [48, 30, 52, 22, 16, 16, 10, 48]
for i, width in enumerate(col_widths_1, 1):
    ws.column_dimensions[get_column_letter(i)].width = width
ws.row_dimensions[4].height = 28


# ============================================================
# ABA 2: RESUMO POR PERFIL
# ============================================================
ws2 = wb.create_sheet("Resumo por Perfil")

ws2.merge_cells('A1:D1')
ws2['A1'] = "Resumo de Horas por Perfil de Analista"
ws2['A1'].font = title_font

headers2 = ["Perfil do Analista", "Horas Totais", "Semanas Alocadas", "Custo Estimado (ref)"]
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

profiles = {}
for item in data:
    if item[3] and item[4]:
        profile = item[3]
        hours = item[4]
        profiles[profile] = profiles.get(profile, 0) + hours

profile_data = [
    ["Arquiteto Cloud / SRE", profiles.get("Arquiteto Cloud", 0), "", "Planejamento, VPN, integracao Avaya, documentacao"],
    ["Cloud Engineer / DevOps", profiles.get("Cloud Engineer", 0), "", "Provisionamento Azure, IaC, networking, seguranca"],
    ["Desenvolvedor Backend (Senior)", profiles.get("Desenvolvedor Backend", 0), "", "Codigo ASR Orchestrator, Speech SDK, Custom Model"],
    ["QA / Tester", profiles.get("QA / Tester", 0), "", "Testes E2E, latencia, resiliencia, sotaques"],
]

for row_idx, row_data in enumerate(profile_data, 4):
    hours = row_data[1]
    weeks = round(hours / 40, 1)
    row_data[2] = f"~{weeks} semanas"
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = wrap_align
        if col_idx == 2:
            cell.alignment = center_align

total_row2 = 4 + len(profile_data)
ws2.cell(row=total_row2, column=1, value="TOTAL").font = total_font
ws2.cell(row=total_row2, column=1).fill = subtotal_fill
ws2.cell(row=total_row2, column=1).border = thin_border
cell = ws2.cell(row=total_row2, column=2, value=total_hours)
cell.font = total_font
cell.fill = subtotal_fill
cell.alignment = center_align
cell.border = thin_border
weeks_total = round(total_hours / 40, 1)
ws2.cell(row=total_row2, column=3, value=f"~{weeks_total} semanas").font = total_font
ws2.cell(row=total_row2, column=3).fill = subtotal_fill
ws2.cell(row=total_row2, column=3).border = thin_border
ws2.cell(row=total_row2, column=4).fill = subtotal_fill
ws2.cell(row=total_row2, column=4).border = thin_border

info_row = total_row2 + 2
ws2.cell(row=info_row, column=1, value="Premissas:").font = Font(bold=True)
ws2.cell(row=info_row + 1, column=1, value="- Semana de 40h (8h/dia x 5 dias)")
ws2.cell(row=info_row + 2, column=1, value="- Atividades podem ser paralelas entre perfis diferentes")
ws2.cell(row=info_row + 3, column=1, value="- Duracao total estimada: 6-7 semanas (com paralelismo)")
ws2.cell(row=info_row + 4, column=1, value="- Dependencia critica: informacoes tecnicas da Atento (IPs, Avaya, firewall)")
ws2.cell(row=info_row + 5, column=1, value="- Custom Speech Model pode exigir iteracoes adicionais se acuracia < 85%")

col_widths_2 = [35, 16, 20, 55]
for i, width in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = width


# ============================================================
# ABA 3: RESUMO POR RECURSO AZURE
# ============================================================
ws3 = wb.create_sheet("Resumo por Recurso Azure")

ws3.merge_cells('A1:D1')
ws3['A1'] = "Horas de Configuracao por Recurso Azure"
ws3['A1'].font = title_font

headers3 = ["Recurso Azure", "Horas Config.", "Horas Desenvolvimento", "Horas Total"]
for col, header in enumerate(headers3, 1):
    cell = ws3.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

resource_data = [
    ["Planejamento / Design Geral", 18, 0, 18],
    ["Azure Subscription + Resource Groups", 2, 0, 2],
    ["VNet + Subnets + NSGs + Route Tables", 8, 0, 8],
    ["VPN Gateway (S2S + BGP + testes)", 16, 0, 16],
    ["Key Vault + Managed Identity", 3, 0, 3],
    ["Redis Cache + Private Endpoint", 3, 0, 3],
    ["Private Endpoints (todos) + DNS", 3, 0, 3],
    ["API Management Developer (VNet injection)", 5, 9, 14],
    ["Virtual Machine D2s v3 (setup + hardening)", 9, 0, 9],
    ["Speech Services + Custom Model", 2, 26, 28],
    ["ASR Orchestrator (codigo da aplicacao)", 0, 50, 50],
    ["Azure Monitor + App Insights + Alertas", 6, 0, 6],
    ["Integracao Avaya (com equipe Atento)", 0, 14, 14],
    ["Testes E2E + Documentacao", 0, 26, 26],
]

for row_idx, row_data in enumerate(resource_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_align if col_idx > 1 else wrap_align

total_row3 = 4 + len(resource_data)
ws3.cell(row=total_row3, column=1, value="TOTAL").font = total_font
ws3.cell(row=total_row3, column=1).fill = subtotal_fill
ws3.cell(row=total_row3, column=1).border = thin_border

config_total = sum(r[1] for r in resource_data)
dev_total = sum(r[2] for r in resource_data)
all_total = sum(r[3] for r in resource_data)

for col_idx, val in enumerate([config_total, dev_total, all_total], 2):
    cell = ws3.cell(row=total_row3, column=col_idx, value=val)
    cell.font = total_font
    cell.fill = subtotal_fill
    cell.alignment = center_align
    cell.border = thin_border

info_row3 = total_row3 + 2
ws3.cell(row=info_row3, column=1, value="Config. = provisionamento e configuracao de infra Azure").font = Font(italic=True, size=9)
ws3.cell(row=info_row3 + 1, column=1, value="Desenvolvimento = codigo, integracao, treinamento ML, testes").font = Font(italic=True, size=9)

col_widths_3 = [42, 18, 22, 16]
for i, width in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = width


# ============================================================
# ABA 4: TIMELINE (GANTT SIMPLIFICADO)
# ============================================================
ws4 = wb.create_sheet("Timeline Semanal")

ws4.merge_cells('A1:I1')
ws4['A1'] = "Timeline de Execucao - MVP PoC (7 semanas)"
ws4['A1'].font = title_font

gantt_headers = ["Fase / Atividade", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "Perfil Principal"]
for col, header in enumerate(gantt_headers, 1):
    cell = ws4.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

active_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
active_font = Font(bold=True, color="FFFFFF", size=9)
empty_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

gantt_data = [
    ["1. Planejamento e Design", [1], "Arquiteto Cloud"],
    ["2. Subscription + VNet + NSGs", [2], "Cloud Engineer"],
    ["3. Key Vault + Secrets", [2], "Cloud Engineer"],
    ["4. VPN Gateway + Tunel S2S", [2, 3], "Cloud Engineer"],
    ["5. Teste conectividade VPN", [3], "Cloud Engineer + Atento"],
    ["6. Speech Services + PE", [2, 3], "Cloud Engineer"],
    ["7. Redis Cache + PE", [3], "Cloud Engineer"],
    ["8. API Management Developer", [3, 4], "Cloud Engineer"],
    ["9. VM D2s v3 (setup + hardening)", [3], "Cloud Engineer"],
    ["10. Custom Speech Model (treino)", [2, 3, 4, 5], "Desenvolvedor Backend"],
    ["11. ASR Orchestrator (codigo)", [3, 4, 5], "Desenvolvedor Backend"],
    ["12. APIM Policies + APIs", [4], "Desenvolvedor Backend"],
    ["13. Observabilidade (Monitor/Alerts)", [4, 5], "Cloud Engineer"],
    ["14. Integracao Avaya", [5, 6], "Arquiteto + Atento"],
    ["15. Testes E2E + Validacao", [6, 7], "QA / Tester"],
    ["16. Documentacao + Handover", [7], "Arquiteto Cloud"],
]

for row_idx, (activity, weeks, profile) in enumerate(gantt_data, 4):
    ws4.cell(row=row_idx, column=1, value=activity).border = thin_border
    ws4.cell(row=row_idx, column=1).alignment = wrap_align

    for w in range(1, 8):
        cell = ws4.cell(row=row_idx, column=w + 1)
        cell.border = thin_border
        cell.alignment = center_align
        if w in weeks:
            cell.fill = active_fill
            cell.font = active_font
            cell.value = "X"
        else:
            cell.fill = empty_fill

    ws4.cell(row=row_idx, column=9, value=profile).border = thin_border
    ws4.cell(row=row_idx, column=9).alignment = wrap_align

col_widths_4 = [38, 6, 6, 6, 6, 6, 6, 6, 28]
for i, width in enumerate(col_widths_4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = width


# ============================================================
# SALVAR
# ============================================================
output_path = os.path.join(
    r'c:\01. Foursys\06. BMAD Cursor\docs\atento\Calculadora',
    '[Atento] Cronograma MVP PoC - Horas por Atividade.xlsx'
)
wb.save(output_path)
print(f"Excel gerado: {output_path}")
print(f"\nResumo:")
print(f"  Total de horas: {total_hours}h")
print(f"  Duracao estimada: ~7 semanas (com paralelismo)")
print(f"  Config infra: {config_total}h | Desenvolvimento: {dev_total}h")
print(f"\nPor perfil:")
for p, h in profiles.items():
    print(f"  {p}: {h}h (~{round(h/40, 1)} semanas)")
print(f"\nAbas criadas:")
print(f"  1. Cronograma por Atividade (detalhado)")
print(f"  2. Resumo por Perfil")
print(f"  3. Resumo por Recurso Azure")
print(f"  4. Timeline Semanal (Gantt)")
