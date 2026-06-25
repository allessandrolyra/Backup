import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

folder = r'c:\01. Foursys\06. BMAD Cursor\docs\atento\Calculadora'
files = os.listdir(folder)
poc_file = None
for f in files:
    if 'v.01' in f.lower() and not f.startswith('~'):
        poc_file = f
        break

full_path = os.path.join(folder, poc_file)
wb_src = openpyxl.load_workbook(full_path, data_only=True)
ws_src = wb_src.active

rows_data = []
for row in ws_src.iter_rows(min_row=4, max_row=11, values_only=False):
    service_cat = row[0].value
    service_type = row[1].value
    custom_name = row[2].value
    description = row[4].value
    cost = row[5].value
    if service_cat and cost:
        rows_data.append({
            'category': service_cat,
            'type': service_type,
            'name': custom_name,
            'description': description,
            'cost': cost
        })

print(f"Registros extraidos: {len(rows_data)}")
for r in rows_data:
    print(f"  {r['type']}: ${r['cost']}")

wb = Workbook()
ws = wb.active
ws.title = "Estimativa MVP PoC"

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
total_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
title_font = Font(bold=True, size=14)
subtitle_font = Font(bold=True, size=11, color="2F5496")

TAXA_BRL = 5.80

ws.merge_cells('A1:E1')
ws['A1'] = "Atento ASR Cloud - Estimativa MVP 1 PoC"
ws['A1'].font = title_font

ws.merge_cells('A2:E2')
ws['A2'] = "Fonte: Azure Pricing Calculator | Maio 2026 | Brazil South"
ws['A2'].font = subtitle_font

headers = [
    "Nome do Servico",
    "Tipo de Servico",
    "Quantidade de Consumo Mensal",
    "Valor Estimado Mensal (USD)",
    "Valor Estimado Mensal (BRL)"
]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

formatted_data = [
    ["VPN Gateway VpnGw1AZ", "Networking", "1 gateway, 730h, 1 tunel S2S, 100GB transfer"],
    ["Azure Speech Services (STT Real-Time)", "AI + Machine Learning", "15 segundos por ligacao, ~120h audio/mes + 1 Custom endpoint"],
    ["API Management - Developer", "Web / API Gateway", "1 unidade, 730h (VNet injection habilitado)"],
    ["Azure Cache for Redis - Standard C0", "Databases / Cache", "1 instancia Standard C0 (250MB), 730h"],
    ["Azure Monitor + App Insights + Log Analytics", "DevOps / Observabilidade", "Ingestao ~0.2 GB/dia logs, metricas basicas"],
    ["Key Vault", "Seguranca", "1 vault, ~100 operacoes, ~100 operacoes avancadas"],
    ["Virtual Machine D2s v3 (2 vCPU, 8GB)", "Compute", "1 VM Linux, 730h (Pay as you go)"],
    ["Private Endpoints (Speech + Redis)", "Networking / Seguranca", "2 endpoints, 730h, 100GB inbound + 100GB outbound"],
]

costs = [float(r['cost']) for r in rows_data]

for row_idx, (row_data, cost) in enumerate(zip(formatted_data, costs), 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

    cell = ws.cell(row=row_idx, column=4, value=cost)
    cell.border = thin_border
    cell.number_format = '$ #,##0.00'
    cell.alignment = Alignment(vertical='center', horizontal='right')

    cell = ws.cell(row=row_idx, column=5, value=cost * TAXA_BRL)
    cell.border = thin_border
    cell.number_format = 'R$ #,##0.00'
    cell.alignment = Alignment(vertical='center', horizontal='right')

total_row = 5 + len(formatted_data)
ws.cell(row=total_row, column=1, value="TOTAL MENSAL").font = total_font
ws.cell(row=total_row, column=1).fill = total_fill
ws.cell(row=total_row, column=1).border = thin_border

for col in range(2, 4):
    ws.cell(row=total_row, column=col).fill = total_fill
    ws.cell(row=total_row, column=col).border = thin_border

total_usd = sum(costs)
cell = ws.cell(row=total_row, column=4, value=total_usd)
cell.font = total_font
cell.fill = total_fill
cell.number_format = '$ #,##0.00'
cell.border = thin_border
cell.alignment = Alignment(horizontal='right')

cell = ws.cell(row=total_row, column=5, value=total_usd * TAXA_BRL)
cell.font = total_font
cell.fill = total_fill
cell.number_format = 'R$ #,##0.00'
cell.border = thin_border
cell.alignment = Alignment(horizontal='right')

info_row = total_row + 2
ws.cell(row=info_row, column=1, value="Regiao Azure:").font = Font(bold=True)
ws.cell(row=info_row, column=2, value="Brazil South (Sao Paulo)")
ws.cell(row=info_row + 1, column=1, value="Taxa de conversao:").font = Font(bold=True)
ws.cell(row=info_row + 1, column=2, value="USD 1.00 = BRL 5.80 (referencia)")
ws.cell(row=info_row + 2, column=1, value="Cenario:").font = Font(bold=True)
ws.cell(row=info_row + 2, column=2, value="PoC - Prova de Conceito (volume reduzido de testes)")
ws.cell(row=info_row + 3, column=1, value="Data da estimativa:").font = Font(bold=True)
ws.cell(row=info_row + 3, column=2, value="19/05/2026")
ws.cell(row=info_row + 4, column=1, value="Observacao:").font = Font(bold=True)
ws.cell(row=info_row + 4, column=2, value="Valores referencia - confirmar com Azure Pricing Calculator no momento da contratacao")

col_widths = [44, 30, 58, 28, 28]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

ws.row_dimensions[4].height = 30
for r in range(5, total_row + 1):
    ws.row_dimensions[r].height = 22

output_path = os.path.join(folder, "[Atento] Estimativa MVP 1 POC - Formatada.xlsx")
wb.save(output_path)
print(f"\nExcel gerado: {output_path}")
print(f"Total: ${total_usd:.2f} USD | R$ {total_usd * TAXA_BRL:.2f} BRL")
