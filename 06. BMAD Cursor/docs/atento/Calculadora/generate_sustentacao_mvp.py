import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
hdr_font = Font(bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
cat_font = Font(bold=True, color="FFFFFF", size=10)
cat_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
sub_font = Font(bold=True, size=10)
sub_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
tot_font = Font(bold=True, size=11, color="FFFFFF")
tot_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
title_font = Font(bold=True, size=14)
title2_font = Font(bold=True, size=11, color="2F5496")
green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
wrap = Alignment(vertical='center', wrap_text=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)


def set_headers(ws, row, headers, widths=None):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = thin
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# ABA 1: RACIONAL DE SUSTENTACAO - 40h/mes
# ============================================================
ws1 = wb.active
ws1.title = "Racional 40h Sustentacao"

ws1.merge_cells('A1:G1')
ws1['A1'] = "Racional de Sustentacao - Ambiente MVP1 ASR Atento"
ws1['A1'].font = title_font

ws1.merge_cells('A2:G2')
ws1['A2'] = "40 horas mensais | Escopo de atividades por categoria | Perfil: Cloud Engineer / SRE"
ws1['A2'].font = title2_font

headers = [
    "Categoria", "Atividade",
    "Frequencia", "Horas/Execucao", "Execucoes/Mes",
    "Horas/Mes", "Recurso Azure Relacionado"
]
set_headers(ws1, 4, headers, [28, 52, 16, 16, 16, 12, 34])

data = [
    # MONITORAMENTO E OBSERVABILIDADE
    ["MONITORAMENTO E OBSERVABILIDADE", None, None, None, None, None, None],
    [None, "Analise diaria de dashboards (latencia E2E, taxa ASR, erros, throughput)",
     "Diaria", 0.25, 22, None, "Azure Monitor + App Insights"],
    [None, "Verificacao de saude do tunel VPN e revisao de alertas disparados",
     "Diaria", 0.25, 22, None, "VPN Gateway + Monitor Alerts"],
    [None, "Analise semanal de logs (KQL) para tendencias e anomalias",
     "Semanal", 0.75, 4, None, "Log Analytics Workspace"],

    # GESTAO DE INCIDENTES
    ["GESTAO DE INCIDENTES", None, None, None, None, None, None],
    [None, "Investigacao e resolucao de incidentes de producao (P1/P2)",
     "Sob demanda", 2.0, 2, None, "Todos os recursos"],
    [None, "Troubleshooting VPN, Speech Services ou Redis",
     "Sob demanda", 1.5, 1, None, "VPN GW + Speech + Redis"],
    [None, "Restart/recovery de VM ou servicos (se necessario)",
     "Sob demanda", 0.5, 1, None, "VM D2s v3"],

    # MANUTENCAO PREVENTIVA
    ["MANUTENCAO PREVENTIVA", None, None, None, None, None, None],
    [None, "Aplicacao de patches de seguranca no OS da VM (Linux updates)",
     "Mensal", 1.5, 1, None, "VM D2s v3"],
    [None, "Rotacao de secrets e certificados no Key Vault",
     "Mensal", 0.5, 1, None, "Key Vault"],
    [None, "Verificacao de expiracao de certificados VPN (IKE/IPSec)",
     "Mensal", 0.5, 1, None, "VPN Gateway"],
    [None, "Review de NSG rules e firewall (auditoria de seguranca)",
     "Mensal", 1.0, 1, None, "NSGs + VNet"],
    [None, "Limpeza de logs antigos e revisao de retencao",
     "Mensal", 0.5, 1, None, "Log Analytics"],

    # PERFORMANCE E CAPACIDADE
    ["PERFORMANCE E CAPACIDADE", None, None, None, None, None, None],
    [None, "Analise de metricas de capacidade (CPU, memoria, conexoes Redis)",
     "Quinzenal", 0.5, 2, None, "VM + Redis + App Insights"],
    [None, "Analise de custos Azure e otimizacao (Azure Advisor)",
     "Mensal", 1.0, 1, None, "Azure Cost Management"],
    [None, "Avaliacao de acuracia do Custom Speech Model (WER mensal)",
     "Mensal", 1.0, 1, None, "Custom Speech Model"],

    # SUPORTE AO CLIENTE
    ["SUPORTE AO CLIENTE (ATENTO)", None, None, None, None, None, None],
    [None, "Reuniao mensal de status e review de SLA com Atento",
     "Mensal", 1.0, 1, None, "Geral"],
    [None, "Elaboracao de relatorio mensal (uptime, chamadas, latencia, acuracia)",
     "Mensal", 1.5, 1, None, "App Insights + Monitor"],
    [None, "Atendimento a chamados/duvidas tecnicas do cliente",
     "Sob demanda", 0.5, 2, None, "Geral"],

    # EVOLUCAO E MELHORIA CONTINUA
    ["EVOLUCAO E MELHORIA CONTINUA", None, None, None, None, None, None],
    [None, "Ajuste fino de alertas (reduzir falso-positivo, adicionar novos)",
     "Mensal", 1.0, 1, None, "Azure Monitor Alerts"],
    [None, "Atualizacao de runbook operacional e documentacao",
     "Mensal", 1.0, 1, None, "Documentacao"],
    [None, "Deploy de hotfixes ou ajustes na aplicacao ASR Orchestrator",
     "Sob demanda", 1.5, 1, None, "VM + App"],
    [None, "Backup de configuracoes (APIM policies, NSG rules, scripts)",
     "Mensal", 0.5, 1, None, "APIM + NSGs + Scripts"],
]

row_idx = 5
total_hours = 0
cat_totals = {}
current_cat = None

for item in data:
    if item[0] and item[1] is None:
        current_cat = item[0]
        cat_totals[current_cat] = 0
        ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
        c = ws1.cell(row=row_idx, column=1, value=item[0])
        c.font = cat_font
        c.fill = cat_fill
        c.alignment = Alignment(vertical='center')
        c.border = thin
        for col in range(2, 8):
            ws1.cell(row=row_idx, column=col).border = thin
        row_idx += 1
    else:
        hours = round(item[3] * item[4], 1)
        total_hours += hours
        if current_cat:
            cat_totals[current_cat] += hours

        values = [
            "", item[1], item[2], item[3], item[4], hours, item[6]
        ]
        for col_idx, value in enumerate(values, 1):
            c = ws1.cell(row=row_idx, column=col_idx, value=value)
            c.border = thin
            c.alignment = wrap
            if col_idx in (4, 5, 6):
                c.alignment = center
            if col_idx == 6:
                c.font = Font(bold=True)
        row_idx += 1

    if item[0] and item[1] is None and current_cat and current_cat != item[0]:
        pass

# Subtotals by category
row_idx += 1
ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
ws1.cell(row=row_idx, column=1, value="RESUMO POR CATEGORIA").font = Font(bold=True, size=11)
row_idx += 1

set_headers(ws1, row_idx, ["Categoria", "", "", "", "", "Horas/Mes", "%"], 
            [28, 52, 16, 16, 16, 12, 34])
row_idx += 1

for cat, hours in cat_totals.items():
    pct = hours / total_hours if total_hours > 0 else 0
    ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
    c = ws1.cell(row=row_idx, column=1, value=cat)
    c.border = thin
    c.alignment = wrap
    for col in range(2, 6):
        ws1.cell(row=row_idx, column=col).border = thin
    c = ws1.cell(row=row_idx, column=6, value=hours)
    c.border = thin
    c.alignment = center
    c.font = sub_font
    c = ws1.cell(row=row_idx, column=7, value=pct)
    c.border = thin
    c.alignment = center
    c.number_format = '0%'
    row_idx += 1

# TOTAL
ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
c = ws1.cell(row=row_idx, column=1, value="TOTAL MENSAL")
c.font = tot_font
c.fill = tot_fill
c.border = thin
for col in range(2, 6):
    ws1.cell(row=row_idx, column=col).fill = tot_fill
    ws1.cell(row=row_idx, column=col).border = thin

c = ws1.cell(row=row_idx, column=6, value=total_hours)
c.font = tot_font
c.fill = tot_fill
c.alignment = center
c.border = thin

c = ws1.cell(row=row_idx, column=7, value=1.0)
c.font = tot_font
c.fill = tot_fill
c.alignment = center
c.number_format = '0%'
c.border = thin


# ============================================================
# ABA 2: ESCOPO E EXCLUSOES
# ============================================================
ws2 = wb.create_sheet("Escopo e Exclusoes")

ws2.merge_cells('A1:C1')
ws2['A1'] = "Escopo de Sustentacao - Incluso vs Excluso"
ws2['A1'].font = title_font

set_headers(ws2, 3, ["Item", "Incluso no Pacote 40h", "Observacao"], [50, 20, 50])

escopo = [
    # INCLUSO
    ["INCLUSO NO PACOTE", None, None],
    ["Monitoramento proativo diario (dashboards, alertas, VPN)", "SIM", "Segunda a sexta, horario comercial (8h-18h)"],
    ["Investigacao e resolucao de incidentes P1/P2", "SIM", "Tempo de resposta: P1=1h, P2=4h (horario comercial)"],
    ["Patching mensal de seguranca (OS Linux da VM)", "SIM", "Janela de manutencao acordada com Atento"],
    ["Rotacao de secrets e certificados", "SIM", "Key Vault + VPN certificates"],
    ["Troubleshooting de VPN, Speech Services, Redis", "SIM", "Ate resolucao ou escalacao"],
    ["Relatorio mensal de operacao (SLA, metricas, custos)", "SIM", "Entregue ate o 5o dia util do mes seguinte"],
    ["Reuniao mensal de status com Atento", "SIM", "1h-1.5h, pauta padrao"],
    ["Deploy de hotfixes e ajustes na aplicacao ASR", "SIM", "Correcoes de bugs, ajustes de config"],
    ["Auditoria mensal de seguranca (NSGs, firewall rules)", "SIM", "Review e documentacao"],
    ["Analise de custos e otimizacao Azure", "SIM", "Azure Advisor + Cost Management"],
    ["Avaliacao mensal de acuracia do Custom Speech Model", "SIM", "Medicao de WER/CER"],
    ["Atualizacao de documentacao operacional (runbook)", "SIM", "Sempre que houver mudanca"],
    ["Backup de configuracoes (APIM, NSGs, scripts)", "SIM", "Versionado em repositorio"],

    # EXCLUSO
    ["EXCLUSO DO PACOTE (cobrado a parte)", None, None],
    ["Desenvolvimento de novas features no ASR Orchestrator", "NAO", "Requer escopo separado + aprovacao"],
    ["Re-treinamento completo do Custom Speech Model", "NAO", "Projeto separado (~20-40h)"],
    ["Migracao de infra (ex: VM para App Service, APIM upgrade)", "NAO", "Projeto de evolucao arquitetural"],
    ["Suporte 24x7 / plantao fora de horario comercial", "NAO", "Disponivel como addon: +20h/mes"],
    ["Disaster Recovery (failover para regiao secundaria)", "NAO", "Requer infra adicional + projeto"],
    ["Load testing e performance tuning", "NAO", "Sob demanda, escopo separado"],
    ["Integracao com novos sistemas da Atento", "NAO", "Projeto separado"],
    ["Atendimento a incidentes causados por mudancas do cliente", "NAO", "Cobrado por hora adicional"],
]

row_idx = 4
for item in escopo:
    if item[1] is None:
        ws2.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
        c = ws2.cell(row=row_idx, column=1, value=item[0])
        c.font = cat_font
        c.fill = cat_fill
        c.border = thin
        for col in range(2, 4):
            ws2.cell(row=row_idx, column=col).border = thin
    else:
        for col_idx, value in enumerate(item, 1):
            c = ws2.cell(row=row_idx, column=col_idx, value=value)
            c.border = thin
            c.alignment = wrap
            if col_idx == 2:
                c.alignment = center
                if value == "SIM":
                    c.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    c.font = Font(bold=True, color="006100")
                else:
                    c.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    c.font = Font(bold=True, color="9C0006")
    row_idx += 1

col_widths_2 = [52, 22, 52]
for i, w in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# ABA 3: SLA E NIVEIS DE SERVICO
# ============================================================
ws3 = wb.create_sheet("SLA e Niveis de Servico")

ws3.merge_cells('A1:E1')
ws3['A1'] = "Niveis de Servico (SLA) - Sustentacao MVP1"
ws3['A1'].font = title_font

set_headers(ws3, 3, [
    "Prioridade", "Descricao", "Tempo de Resposta",
    "Tempo de Resolucao", "Exemplos"
], [16, 40, 20, 22, 48])

sla_data = [
    ["P1 - Critico", "Servico ASR totalmente indisponivel", "1 hora",
     "4 horas", "VPN down, VM parada, Speech Services fora, APIM indisponivel"],
    ["P2 - Alto", "Degradacao significativa de performance ou funcionalidade", "4 horas",
     "8 horas (1 dia util)", "Latencia > 2s, taxa ASR < 70%, Redis indisponivel"],
    ["P3 - Medio", "Problema parcial sem impacto critico na operacao", "8 horas",
     "3 dias uteis", "Alertas intermitentes, logs com erros nao-bloqueantes, dashboard fora"],
    ["P4 - Baixo", "Solicitacao de informacao, melhoria ou ajuste menor", "24 horas",
     "5 dias uteis", "Ajuste de alerta, pedido de relatorio, duvida tecnica"],
]

for i, row_data in enumerate(sla_data, 4):
    for col, v in enumerate(row_data, 1):
        c = ws3.cell(row=i, column=col, value=v)
        c.border = thin
        c.alignment = wrap

r = 4 + len(sla_data) + 2
ws3.cell(row=r, column=1, value="PREMISSAS DO SLA:").font = Font(bold=True, size=11)
premissas = [
    "- Horario de cobertura: Segunda a Sexta, 08h-18h (horario de Brasilia)",
    "- Chamados abertos fora do horario serao triados no proximo dia util",
    "- Meta de disponibilidade do servico ASR: 99.5% mensal",
    "- Credito de SLA: 5% da mensalidade por cada 0.1% abaixo de 99.5%",
    "- Incidentes P1 fora do horario: contato de emergencia disponivel (sem SLA formal)",
    "- Maximo de 2 incidentes P1 simultaneos cobertos pelo pacote",
    "- Excedente de horas (acima de 40h): cobrado a R$ 250/hora adicional",
]
for i, p in enumerate(premissas):
    ws3.cell(row=r + 1 + i, column=1, value=p)

col_widths_3 = [18, 42, 22, 24, 50]
for i, w in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# ABA 4: DISTRIBUICAO MENSAL VISUAL
# ============================================================
ws4 = wb.create_sheet("Distribuicao Mensal")

ws4.merge_cells('A1:F1')
ws4['A1'] = "Distribuicao das 40 Horas ao Longo do Mes"
ws4['A1'].font = title_font

set_headers(ws4, 3, [
    "Tipo de Atividade", "Frequencia", "Horas/Mes",
    "S1 (D1-D7)", "S2 (D8-D14)", "S3 (D15-D21)", "S4 (D22-D30)"
], [38, 16, 12, 14, 14, 14, 14])

active_fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
active_font = Font(bold=True, color="FFFFFF", size=9)
light_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
empty_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

dist_data = [
    ["Monitoramento diario (dashboards, alertas, VPN)", "Diaria", 22.0, "X", "X", "X", "X"],
    ["Analise semanal de logs (KQL)", "Semanal", 4.0, "X", "X", "X", "X"],
    ["Analise de capacidade (CPU, mem, Redis)", "Semanal", 2.0, "X", "X", "X", "X"],
    ["Patching de seguranca (OS Linux)", "Mensal", 1.5, "", "", "", "X"],
    ["Rotacao secrets/certificados", "Mensal", 0.5, "", "", "", "X"],
    ["Verificacao certificados VPN", "Mensal", 0.5, "", "X", "", ""],
    ["Auditoria NSG/firewall", "Mensal", 1.0, "", "", "X", ""],
    ["Limpeza de logs", "Mensal", 0.5, "", "", "", "X"],
    ["Analise de custos Azure", "Mensal", 1.0, "X", "", "", ""],
    ["Avaliacao acuracia Speech Model", "Mensal", 1.5, "", "", "X", ""],
    ["Relatorio mensal + Reuniao status", "Mensal", 3.5, "X", "", "", ""],
    ["Atendimento chamados cliente", "Sob demanda", 1.0, "~", "~", "~", "~"],
    ["Deploy hotfixes / ajustes app", "Sob demanda", 1.5, "~", "~", "~", "~"],
    ["Ajuste de alertas + documentacao", "Mensal", 2.0, "", "X", "", ""],
    ["Backup de configuracoes", "Mensal", 0.5, "", "", "", "X"],
    ["Buffer incidentes / demandas nao previstas", "Sob demanda", 5.5, "~", "~", "~", "~"],
]

# Verify total
dist_total = sum(d[2] for d in dist_data)

for i, row_data in enumerate(dist_data, 4):
    for col, v in enumerate(row_data, 1):
        c = ws4.cell(row=i, column=col, value=v)
        c.border = thin
        c.alignment = center if col >= 3 else wrap
        if col >= 4:
            if v == "X":
                c.fill = active_fill
                c.font = active_font
            elif v == "~":
                c.fill = light_fill
                c.font = Font(italic=True, size=9)
            else:
                c.fill = empty_fill
        if col == 3:
            c.font = Font(bold=True)

total_row = 4 + len(dist_data)
ws4.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
c = ws4.cell(row=total_row, column=1, value="TOTAL")
c.font = tot_font
c.fill = tot_fill
c.border = thin
ws4.cell(row=total_row, column=2).fill = tot_fill
ws4.cell(row=total_row, column=2).border = thin

c = ws4.cell(row=total_row, column=3, value=dist_total)
c.font = tot_font
c.fill = tot_fill
c.alignment = center
c.border = thin

for col in range(4, 8):
    ws4.cell(row=total_row, column=col).fill = tot_fill
    ws4.cell(row=total_row, column=col).border = thin

ws4.column_dimensions[get_column_letter(7)].width = 14

r = total_row + 2
ws4.cell(row=r, column=1, value="Legenda:").font = Font(bold=True)
c = ws4.cell(row=r+1, column=1, value="X = Atividade planejada nesta semana")
c = ws4.cell(row=r+2, column=1, value="~ = Sob demanda (pode ocorrer em qualquer semana)")
ws4.cell(row=r+3, column=1, value=f"Total calculado: {dist_total}h (meta: 40h)")


# ============================================================
# SALVAR
# ============================================================
output = os.path.join(
    r'c:\01. Foursys\06. BMAD Cursor\docs\atento\Calculadora',
    '[Atento] Racional Sustentacao MVP1 - 40h.xlsx'
)
wb.save(output)

print(f"Excel gerado: {output}")
print(f"\nTotal de horas: {total_hours}h")
print(f"\nDistribuicao por categoria:")
for cat, hours in cat_totals.items():
    pct = hours / total_hours * 100
    print(f"  {cat}: {hours}h ({pct:.0f}%)")
print(f"\nAbas:")
print(f"  1. Racional 40h Sustentacao (atividades detalhadas)")
print(f"  2. Escopo e Exclusoes (o que esta incluso e o que nao)")
print(f"  3. SLA e Niveis de Servico (P1-P4 com tempos)")
print(f"  4. Distribuicao Mensal (visual semanal)")
