import subprocess
import sys

try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
total_font = Font(bold=True, size=11)
total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
title_font = Font(bold=True, size=14)
subtitle_font = Font(bold=True, size=11, color="2F5496")

TAXA_BRL = 5.80

# --- CENÁRIO B: MVP PRODUÇÃO (RECOMENDADO) ---
ws = wb.active
ws.title = "MVP Produção (Recomendado)"

ws.merge_cells('A1:E1')
ws['A1'] = "Atento ASR Cloud — Estimativa MVP Produção (10 chamadas/min)"
ws['A1'].font = title_font

ws.merge_cells('A2:E2')
ws['A2'] = "Cenário B — Equilíbrio custo × confiabilidade | Maio 2026"
ws['A2'].font = subtitle_font

headers = ["Nome do Serviço", "Tipo de Serviço", "SKU / Tier", "Qtd Consumo Mensal", "Valor Estimado Mensal (USD)", "Valor Estimado Mensal (BRL)"]

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

data_b = [
    ["VPN Gateway", "Networking", "VpnGw1AZ (Zone-Redundant)", "1 gateway / 100GB transfer", 156.95],
    ["App Service", "Compute", "Premium V3 P0V3 (1 vCPU, 4GB) Linux", "1 instância", 80.00],
    ["API Management", "API Gateway", "Eliminado (middleware no app)", "N/A — implementado no App Service", 0.00],
    ["Azure Speech Services", "AI / Cognitive Services", "S0 — STT Real-Time + Custom Model", "440 horas de áudio + 1 endpoint Custom", 463.65],
    ["Redis Cache", "Data / Cache", "Basic C0 (250MB)", "1 instância", 16.00],
    ["Azure Monitor + App Insights", "Observabilidade", "Pay-per-use (~0.4 GB/dia)", "~12 GB logs/mês + métricas", 100.00],
    ["Key Vault", "Segurança", "Standard", "1 vault / ~1.000 operações", 13.18],
    ["Private Endpoints", "Networking", "PE para App Service, Speech, Redis", "3 endpoints", 21.90],
]

for row_idx, row_data in enumerate(data_b, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx == 5:
            cell.number_format = '$#,##0.00'
    brl_value = row_data[4] * TAXA_BRL
    cell = ws.cell(row=row_idx, column=6, value=brl_value)
    cell.border = thin_border
    cell.number_format = 'R$ #,##0.00'
    cell.alignment = Alignment(vertical='center')

total_row = 5 + len(data_b)
ws.cell(row=total_row, column=1, value="TOTAL MENSAL").font = total_font
ws.cell(row=total_row, column=1).fill = total_fill
ws.cell(row=total_row, column=1).border = thin_border
for col in range(2, 5):
    ws.cell(row=total_row, column=col).fill = total_fill
    ws.cell(row=total_row, column=col).border = thin_border

total_usd = sum(r[4] for r in data_b)
cell = ws.cell(row=total_row, column=5, value=total_usd)
cell.font = total_font
cell.fill = total_fill
cell.number_format = '$#,##0.00'
cell.border = thin_border

cell = ws.cell(row=total_row, column=6, value=total_usd * TAXA_BRL)
cell.font = total_font
cell.fill = total_fill
cell.number_format = 'R$ #,##0.00'
cell.border = thin_border

info_row = total_row + 2
ws.cell(row=info_row, column=1, value="Economia vs estimativa anterior:").font = Font(bold=True)
ws.cell(row=info_row, column=2, value="50% (de $1.712 para $852/mês)")
ws.cell(row=info_row + 1, column=1, value="Taxa de conversão:").font = Font(bold=True)
ws.cell(row=info_row + 1, column=2, value="USD 1.00 = BRL 5.80")
ws.cell(row=info_row + 2, column=1, value="Região Azure:").font = Font(bold=True)
ws.cell(row=info_row + 2, column=2, value="Brazil South (São Paulo)")

col_widths = [30, 25, 38, 40, 28, 28]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# --- CENÁRIO A: ULTRA-MVP (PoC) ---
ws_a = wb.create_sheet("Ultra-MVP (PoC)")

ws_a.merge_cells('A1:E1')
ws_a['A1'] = "Atento ASR Cloud — Estimativa Ultra-MVP / PoC (10 chamadas/min)"
ws_a['A1'].font = title_font

ws_a.merge_cells('A2:E2')
ws_a['A2'] = "Cenário A — Custo mínimo absoluto | Maio 2026"
ws_a['A2'].font = subtitle_font

for col, header in enumerate(headers, 1):
    cell = ws_a.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

data_a = [
    ["VPN Gateway", "Networking", "VpnGw1 (non-AZ)", "1 gateway / 100GB transfer", 138.00],
    ["App Service", "Compute", "Premium V3 P0V3 (1 vCPU, 4GB) Linux", "1 instância", 80.00],
    ["API Management", "API Gateway", "Eliminado (middleware no app)", "N/A", 0.00],
    ["Azure Speech Services", "AI / Cognitive Services", "S0 — STT Real-Time + Custom Model", "440 horas de áudio + 1 endpoint Custom", 463.65],
    ["Redis Cache", "Data / Cache", "Eliminado (in-memory no App Service)", "N/A", 0.00],
    ["Azure Monitor + App Insights", "Observabilidade", "Pay-per-use (~0.2 GB/dia)", "~6 GB logs/mês + métricas", 60.00],
    ["Key Vault", "Segurança", "Standard", "1 vault / ~1.000 operações", 13.18],
    ["Private Endpoint (App Service)", "Networking", "1 PE", "1 endpoint", 7.30],
    ["Private Endpoint (Speech)", "Networking", "1 PE", "1 endpoint", 7.30],
]

for row_idx, row_data in enumerate(data_a, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_a.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx == 5:
            cell.number_format = '$#,##0.00'
    brl_value = row_data[4] * TAXA_BRL
    cell = ws_a.cell(row=row_idx, column=6, value=brl_value)
    cell.border = thin_border
    cell.number_format = 'R$ #,##0.00'
    cell.alignment = Alignment(vertical='center')

total_row_a = 5 + len(data_a)
ws_a.cell(row=total_row_a, column=1, value="TOTAL MENSAL").font = total_font
ws_a.cell(row=total_row_a, column=1).fill = total_fill
ws_a.cell(row=total_row_a, column=1).border = thin_border
for col in range(2, 5):
    ws_a.cell(row=total_row_a, column=col).fill = total_fill
    ws_a.cell(row=total_row_a, column=col).border = thin_border

total_usd_a = sum(r[4] for r in data_a)
cell = ws_a.cell(row=total_row_a, column=5, value=total_usd_a)
cell.font = total_font
cell.fill = total_fill
cell.number_format = '$#,##0.00'
cell.border = thin_border

cell = ws_a.cell(row=total_row_a, column=6, value=total_usd_a * TAXA_BRL)
cell.font = total_font
cell.fill = total_fill
cell.number_format = 'R$ #,##0.00'
cell.border = thin_border

for i, width in enumerate(col_widths, 1):
    ws_a.column_dimensions[get_column_letter(i)].width = width

# --- CENÁRIO 100/min OTIMIZADO ---
ws_100 = wb.create_sheet("100-min (Sem AKS)")

ws_100.merge_cells('A1:E1')
ws_100['A1'] = "Atento ASR Cloud — Estimativa 100 chamadas/min (Sem AKS)"
ws_100['A1'].font = title_font

ws_100.merge_cells('A2:E2')
ws_100['A2'] = "App Service substitui AKS — menor complexidade operacional | Maio 2026"
ws_100['A2'].font = subtitle_font

for col, header in enumerate(headers, 1):
    cell = ws_100.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

data_100 = [
    ["VPN Gateway", "Networking", "VpnGw2AZ (Zone-Redundant)", "1 gateway / 1.2TB transfer", 458.45],
    ["App Service", "Compute", "Premium V3 P1V3 (2 vCPU, 8GB) Linux", "2 instâncias (autoscale)", 248.00],
    ["API Management", "API Gateway", "Eliminado (middleware no app)", "N/A — reintroduzir Standard v2 se necessário", 0.00],
    ["Azure Speech Services", "AI / Cognitive Services", "S0 — STT Real-Time + Custom Model", "4.400 horas de áudio + 1 endpoint Custom", 4423.65],
    ["Redis Cache", "Data / Cache", "Standard C1 (1GB)", "1 instância", 100.74],
    ["Azure Monitor + App Insights", "Observabilidade", "Pay-per-use (~0.4 GB/dia)", "~12 GB logs/mês + métricas", 128.90],
    ["Key Vault", "Segurança", "Standard", "1 vault / ~10.000 operações", 13.18],
    ["Private Endpoints", "Networking", "PE para App, Speech, Redis, KV", "4 endpoints", 29.20],
]

for row_idx, row_data in enumerate(data_100, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_100.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx == 5:
            cell.number_format = '$#,##0.00'
    brl_value = row_data[4] * TAXA_BRL
    cell = ws_100.cell(row=row_idx, column=6, value=brl_value)
    cell.border = thin_border
    cell.number_format = 'R$ #,##0.00'
    cell.alignment = Alignment(vertical='center')

total_row_100 = 5 + len(data_100)
ws_100.cell(row=total_row_100, column=1, value="TOTAL MENSAL").font = total_font
ws_100.cell(row=total_row_100, column=1).fill = total_fill
ws_100.cell(row=total_row_100, column=1).border = thin_border
for col in range(2, 5):
    ws_100.cell(row=total_row_100, column=col).fill = total_fill
    ws_100.cell(row=total_row_100, column=col).border = thin_border

total_usd_100 = sum(r[4] for r in data_100)
cell = ws_100.cell(row=total_row_100, column=5, value=total_usd_100)
cell.font = total_font
cell.fill = total_fill
cell.number_format = '$#,##0.00'
cell.border = thin_border

cell = ws_100.cell(row=total_row_100, column=6, value=total_usd_100 * TAXA_BRL)
cell.font = total_font
cell.fill = total_fill
cell.number_format = 'R$ #,##0.00'
cell.border = thin_border

info_row_100 = total_row_100 + 2
ws_100.cell(row=info_row_100, column=1, value="Economia vs estimativa anterior (com AKS):").font = Font(bold=True)
ws_100.cell(row=info_row_100, column=2, value="11% (de $6.061 para $5.402/mês)")
ws_100.cell(row=info_row_100 + 1, column=1, value="Observação:").font = Font(bold=True)
ws_100.cell(row=info_row_100 + 1, column=2, value="Speech Services representa 82% do custo neste cenário")

for i, width in enumerate(col_widths, 1):
    ws_100.column_dimensions[get_column_letter(i)].width = width

# --- COMPARATIVO ---
ws_comp = wb.create_sheet("Comparativo")

ws_comp.merge_cells('A1:F1')
ws_comp['A1'] = "Comparativo: Estimativa Anterior vs MVP Otimizado"
ws_comp['A1'].font = title_font

comp_headers = ["Cenário", "USD/mês", "BRL/mês", "Economia (%)", "Compute", "APIM"]
for col, header in enumerate(comp_headers, 1):
    cell = ws_comp.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

comp_data = [
    ["10/min — Anterior", 1712.19, 1712.19 * TAXA_BRL, "—", "App Service P1V3", "Standard ($687)"],
    ["10/min — Ultra-MVP (A)", total_usd_a, total_usd_a * TAXA_BRL, "55%", "App Service P0V3", "Eliminado"],
    ["10/min — MVP Produção (B)", total_usd, total_usd * TAXA_BRL, "50%", "App Service P0V3", "Eliminado"],
    ["100/min — Anterior", 6061.19, 6061.19 * TAXA_BRL, "—", "AKS (Reserved 3yr)", "Standard ($687)"],
    ["100/min — Otimizado", total_usd_100, total_usd_100 * TAXA_BRL, "11%", "App Service P1V3 ×2", "Eliminado"],
]

for row_idx, row_data in enumerate(comp_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_comp.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', horizontal='center')
        if col_idx == 2:
            cell.number_format = '$#,##0.00'
        elif col_idx == 3:
            cell.number_format = 'R$ #,##0.00'

comp_widths = [30, 18, 18, 15, 25, 22]
for i, width in enumerate(comp_widths, 1):
    ws_comp.column_dimensions[get_column_letter(i)].width = width

output_path = r"c:\01. Foursys\06. BMAD Cursor\docs\atento\Atento-ASR-Estimativa-MVP-Otimizada.xlsx"
wb.save(output_path)
print(f"Excel gerado com sucesso: {output_path}")
print(f"\nAbas criadas:")
print(f"  1. MVP Produção (Recomendado) — Cenário B: ${total_usd:.2f}/mês")
print(f"  2. Ultra-MVP (PoC) — Cenário A: ${total_usd_a:.2f}/mês")
print(f"  3. 100-min (Sem AKS) — ${total_usd_100:.2f}/mês")
print(f"  4. Comparativo — Resumo geral")
